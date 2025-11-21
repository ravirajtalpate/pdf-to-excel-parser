"""
AI-Powered Document Structuring & Data Extraction
Converts unstructured PDF documents into structured Excel output
"""

import PyPDF2
import pandas as pd
import re
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import anthropic
import os
import json

class PDFToExcelExtractor:
    def __init__(self, pdf_path: str, api_key: str = None):
        """
        Initialize the PDF to Excel extractor
        
        Args:
            pdf_path: Path to input PDF file
            api_key: Anthropic API key (optional, will use env variable if not provided)
        """
        self.pdf_path = pdf_path
        self.api_key = api_key or os.getenv('ANTHROPIC_API_KEY')
        self.client = anthropic.Anthropic(api_key=self.api_key)
        
    def extract_text_from_pdf(self) -> str:
        """Extract all text content from PDF"""
        text = ""
        try:
            with open(self.pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            print(f"Error extracting PDF: {e}")
            raise
        return text
    
    def structure_data_with_ai(self, pdf_text: str) -> List[Dict]:
        """
        Use Claude AI to structure the unstructured PDF text into key-value pairs
        
        Args:
            pdf_text: Raw text extracted from PDF
            
        Returns:
            List of dictionaries with Key, Value, and Comments
        """
        prompt = f"""You are an expert data extraction assistant. Your task is to analyze the following unstructured text from a PDF document and convert it into a structured format.

CRITICAL REQUIREMENTS:
1. Extract ALL information - nothing should be lost or omitted
2. Identify logical key-value relationships in the text
3. Create structured key-value pairs where:
   - **Key**: A clear, descriptive label (e.g., "12th standard pass out year", "Undergraduate degree")
   - **Value**: The core data/fact (e.g., "2007", "B.Tech (Computer Science)")
   - **Comments**: ALL remaining descriptive text, achievements, context, explanations, or elaborations related to that key-value pair
4. Preserve the EXACT original wording from the PDF - do not paraphrase
5. Handle multi-line or complex text structures faithfully

COMMENTS FIELD RULES (VERY IMPORTANT):
- The Comments field should contain ALL additional information beyond the simple key-value pair
- Include descriptions, achievements, rankings, scores, contextual details, explanations
- If there are multiple sentences about a topic, ALL of them go in Comments
- Examples:
  * For education: Include subjects studied, rankings, honors, class performance
  * For certifications: Include scores, years obtained, achievement levels, ratings
  * For skills: Include proficiency levels, specific tools, expertise ratings
- Comments should be substantial and informative, NOT empty unless truly no context exists
- Preserve exact wording from the PDF in Comments

EXAMPLE FORMAT:
If PDF says: "Passed 12th standard in 2007 with 92.50% score. His core subjects included Mathematics, Physics, Chemistry, and Computer Science, demonstrating his early aptitude for technical fields. This was an outstanding achievement."

Extract as:
{{
    "Key": "12th standard pass out year",
    "Value": "2007",
    "Comments": "His core subjects included Mathematics, Physics, Chemistry, and Computer Science, demonstrating his early aptitude for technical fields."
}},
{{
    "Key": "12th overall board score",
    "Value": "92.50%",
    "Comments": "Outstanding achievement"
}}

For educational/professional records, use keys like:
- "12th standard pass out year"
- "12th overall board score"
- "Undergraduate degree"
- "Undergraduate college"
- "Undergraduate year"
- "Undergraduate CGPA"
- "Graduation degree"
- "Graduation college"
- "Graduation year"
- "Graduation CGPA"
- "Certifications 1", "Certifications 2", "Certifications 3", "Certifications 4"
- "Technical Proficiency"

PDF TEXT TO ANALYZE:
{pdf_text}

OUTPUT FORMAT:
Return a JSON array where each element has this structure:
{{
    "Key": "descriptive key name",
    "Value": "extracted value",
    "Comments": "ALL additional context, descriptions, achievements, or notes"
}}

CRITICAL REMINDERS:
- Return ONLY the JSON array, no other text or markdown
- Ensure 100% of the PDF content is captured across all fields (Key, Value, Comments)
- Comments should be rich and detailed, NOT empty
- Maintain original language and phrasing
- Number similar keys (e.g., Certifications 1, Certifications 2, Certifications 3)
"""

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,  # Increased for better comment extraction
                temperature=0,  # More deterministic output
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            
            # Extract the text response
            response_text = message.content[0].text
            
            # Clean up the response - remove markdown code blocks if present
            response_text = response_text.strip()
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            # Parse JSON response
            structured_data = json.loads(response_text)
            
            # Validate and enhance comments
            for item in structured_data:
                # Ensure all required fields exist
                if 'Key' not in item or 'Value' not in item:
                    print(f"Warning: Incomplete item found: {item}")
                    continue
                
                # Ensure Comments field exists
                if 'Comments' not in item:
                    item['Comments'] = ""
                
                # Debug output
                print(f"✓ {item['Key']}: {item['Value'][:50] if len(str(item['Value'])) > 50 else item['Value']}")
                if item['Comments']:
                    print(f"  └─ Comment: {item['Comments'][:100]}...")
            
            return structured_data
            
        except Exception as e:
            print(f"Error in AI structuring: {e}")
            raise
    
    def enhance_comments(self, structured_data: List[Dict], pdf_text: str) -> List[Dict]:
        """
        Second pass: Enhance comments with more context from the PDF
        
        Args:
            structured_data: Initial structured data
            pdf_text: Original PDF text
            
        Returns:
            Enhanced structured data with richer comments
        """
        # Create a summary of extracted data
        data_summary = "\n".join([f"{item['Key']}: {item['Value']}" for item in structured_data])
        
        prompt = f"""You have extracted key-value pairs from a document. Now review the ORIGINAL text and ENHANCE the Comments field for each item.

ORIGINAL PDF TEXT:
{pdf_text}

CURRENTLY EXTRACTED DATA:
{json.dumps(structured_data, indent=2)}

TASK:
For each item in the extracted data, find ALL related contextual information from the original PDF text and add it to the Comments field. 

RULES FOR COMMENTS:
1. Include achievements, rankings, scores, descriptions, explanations
2. Include any elaboration or detail about the key-value pair
3. Use EXACT wording from the original PDF
4. If multiple sentences relate to an item, include ALL of them
5. Comments should be substantial and informative

Return the COMPLETE JSON array with enhanced Comments fields. Keep the same structure:
{{
    "Key": "same as before",
    "Value": "same as before",
    "Comments": "ENHANCED with all contextual information from PDF"
}}

Return ONLY the JSON array, no other text."""

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,
                temperature=0,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            
            response_text = message.content[0].text.strip()
            
            # Clean up response
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            enhanced_data = json.loads(response_text)
            
            print("\n📝 Comments Enhancement Complete:")
            for item in enhanced_data:
                if item.get('Comments'):
                    print(f"✓ Enhanced {item['Key']}")
            
            return enhanced_data
            
        except Exception as e:
            print(f"Warning: Could not enhance comments: {e}")
            return structured_data  # Return original if enhancement fails
    
        """
        Create formatted Excel file from structured data
        
        Args:
            structured_data: List of dictionaries with Key, Value, Comments
            output_path: Path for output Excel file
        """
        # Create DataFrame
        df = pd.DataFrame(structured_data)
        
        # Ensure columns are in the correct order
        df = df[['Key', 'Value', 'Comments']]
        
        # Add row numbers (starting from 23 to match the screenshot)
        df.insert(0, '#', range(23, 23 + len(df)))
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Output', index=False, startrow=1)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Output']
            
            # Style the header row
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header styling
            for cell in worksheet[2]:  # Header is in row 2
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply borders to all data cells
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 5   # #
            worksheet.column_dimensions['B'].width = 30  # Key
            worksheet.column_dimensions['C'].width = 40  # Value
            worksheet.column_dimensions['D'].width = 80  # Comments
            
        print(f"Excel file created successfully: {output_path}")
    
    def process(self, output_path: str = "Output.xlsx"):
        """
        Complete processing pipeline: PDF -> AI Structuring -> Excel
        
        Args:
            output_path: Path for output Excel file
        """
        print("Step 1: Extracting text from PDF...")
        pdf_text = self.extract_text_from_pdf()
        print(f"Extracted {len(pdf_text)} characters from PDF")
        
        print("\nStep 2: Structuring data with AI...")
        structured_data = self.structure_data_with_ai(pdf_text)
        print(f"Extracted {len(structured_data)} key-value pairs")
        
        print("\nStep 3: Enhancing comments with contextual information...")
        structured_data = self.enhance_comments(structured_data, pdf_text)
        
        print("\nStep 4: Creating Excel output...")
        self.create_excel_output(structured_data, output_path)
        print("\n✓ Process completed successfully!")
        
        return structured_data


def main():
    """Main execution function"""
    # Configuration
    PDF_INPUT = "Data Input.pdf"
    EXCEL_OUTPUT = "Output.xlsx"
    
    # Check for API key
    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY environment variable not set!")
        print("Please set it using: export ANTHROPIC_API_KEY='your-api-key'")
        return
    
    # Check if input file exists
    if not os.path.exists(PDF_INPUT):
        print(f"ERROR: Input file '{PDF_INPUT}' not found!")
        return
    
    try:
        # Initialize and process
        extractor = PDFToExcelExtractor(PDF_INPUT, api_key)
        structured_data = extractor.process(EXCEL_OUTPUT)
        
        # Display summary
        print("\n" + "="*60)
        print("EXTRACTION SUMMARY")
        print("="*60)
        print(f"Total records extracted: {len(structured_data)}")
        print(f"Output file: {EXCEL_OUTPUT}")
        print("="*60)
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
