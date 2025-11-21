"""
Streamlit Web Application for AI-Powered PDF to Excel Extraction
"""

import streamlit as st
import PyPDF2
import pandas as pd
import re
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import anthropic
import json
import io
import base64
from datetime import datetime

# Page configuration
st.set_page_config(
    page_title="AI PDF to Excel Extractor",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
    }
    .info-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
    }
    .warning-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        color: #856404;
    }
    </style>
""", unsafe_allow_html=True)


class PDFToExcelExtractor:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.client = anthropic.Anthropic(api_key=self.api_key)
        
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract all text content from uploaded PDF"""
        text = ""
        try:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        except Exception as e:
            st.error(f"Error extracting PDF: {e}")
            raise
        return text
    
    def structure_data_with_ai(self, pdf_text: str) -> List[Dict]:
        """Use Claude AI to structure the unstructured PDF text"""
        prompt = f"""You are an expert data extraction assistant. Your task is to analyze the following unstructured text from a PDF document and convert it into a structured format.

CRITICAL REQUIREMENTS:
1. Extract ALL information - nothing should be lost or omitted
2. Identify logical key-value relationships in the text
3. Create structured key-value pairs where:
   - **Key**: A clear, descriptive label (e.g., "12th standard pass out year", "Undergraduate degree")
   - **Value**: The core data/fact (e.g., "2007", "B.Tech (Computer Science)")
   - **Comments**: ALL remaining descriptive text, achievements, context, explanations, or elaborations related to that key-value pair
4. Preserve the EXACT original wording from the PDF - do not paraphrase
5. Handle multi-line or complex text structures faithfully

COMMENTS FIELD RULES (VERY IMPORTANT):
- The Comments field should contain ALL additional information beyond the simple key-value pair
- Include descriptions, achievements, rankings, scores, contextual details, explanations
- If there are multiple sentences about a topic, ALL of them go in Comments
- Examples:
  * For education: Include subjects studied, rankings, honors, class performance
  * For certifications: Include scores, years obtained, achievement levels, ratings
  * For skills: Include proficiency levels, specific tools, expertise ratings
- Comments should be substantial and informative, NOT empty unless truly no context exists
- Preserve exact wording from the PDF in Comments

For educational/professional records, use keys like:
- "12th standard pass out year"
- "12th overall board score"
- "Undergraduate degree"
- "Undergraduate college"
- "Undergraduate year"
- "Undergraduate CGPA"
- "Graduation degree"
- "Graduation college"
- "Graduation year"
- "Graduation CGPA"
- "Certifications 1", "Certifications 2", "Certifications 3", "Certifications 4"
- "Technical Proficiency"

PDF TEXT TO ANALYZE:
{pdf_text}

OUTPUT FORMAT:
Return a JSON array where each element has this structure:
{{
    "Key": "descriptive key name",
    "Value": "extracted value",
    "Comments": "ALL additional context, descriptions, achievements, or notes"
}}

CRITICAL REMINDERS:
- Return ONLY the JSON array, no other text or markdown
- Ensure 100% of the PDF content is captured across all fields (Key, Value, Comments)
- Comments should be rich and detailed, NOT empty
- Maintain original language and phrasing
- Number similar keys (e.g., Certifications 1, Certifications 2, Certifications 3)
"""

        try:
            with st.spinner("🤖 AI is analyzing and structuring your document..."):
                message = self.client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=8000,
                    temperature=0,
                    messages=[
                        {"role": "user", "content": prompt}
                    ]
                )
            
            response_text = message.content[0].text.strip()
            
            # Clean up the response
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            structured_data = json.loads(response_text)
            
            # Validate data
            for item in structured_data:
                if 'Comments' not in item:
                    item['Comments'] = ""
            
            return structured_data
            
        except Exception as e:
            st.error(f"Error in AI structuring: {e}")
            raise
    
    def enhance_comments(self, structured_data: List[Dict], pdf_text: str) -> List[Dict]:
        """Second pass: Enhance comments with more context"""
        prompt = f"""You have extracted key-value pairs from a document. Now review the ORIGINAL text and ENHANCE the Comments field for each item.

ORIGINAL PDF TEXT:
{pdf_text}

CURRENTLY EXTRACTED DATA:
{json.dumps(structured_data, indent=2)}

TASK:
For each item in the extracted data, find ALL related contextual information from the original PDF text and add it to the Comments field. 

RULES FOR COMMENTS:
1. Include achievements, rankings, scores, descriptions, explanations
2. Include any elaboration or detail about the key-value pair
3. Use EXACT wording from the original PDF
4. If multiple sentences relate to an item, include ALL of them
5. Comments should be substantial and informative

Return the COMPLETE JSON array with enhanced Comments fields. Keep the same structure:
{{
    "Key": "same as before",
    "Value": "same as before",
    "Comments": "ENHANCED with all contextual information from PDF"
}}

Return ONLY the JSON array, no other text."""

        try:
            with st.spinner("✨ Enhancing comments with contextual information..."):
                message = self.client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=8000,
                    temperature=0,
                    messages=[
                        {"role": "user", "content": prompt}
                    ]
                )
            
            response_text = message.content[0].text.strip()
            
            # Clean up response
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            enhanced_data = json.loads(response_text)
            return enhanced_data
            
        except Exception as e:
            st.warning(f"Could not enhance comments: {e}")
            return structured_data
    
    def create_excel_output(self, structured_data: List[Dict]) -> bytes:
        """Create formatted Excel file and return as bytes"""
        df = pd.DataFrame(structured_data)
        df = df[['Key', 'Value', 'Comments']]
        df.insert(0, '#', range(23, 23 + len(df)))
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Output', index=False, startrow=1)
            
            workbook = writer.book
            worksheet = writer.sheets['Output']
            
            # Style the header row
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header styling
            for cell in worksheet[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply borders to all data cells
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 5
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 80
        
        output.seek(0)
        return output.getvalue()


def main():
    # Header
    st.markdown('<div class="main-header">📄 AI-Powered PDF to Excel Extractor</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Transform unstructured documents into structured Excel data using Claude AI</div>', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        api_key = st.text_input(
            "Anthropic API Key",
            type="password",
            help="Enter your Anthropic API key. Get one at https://console.anthropic.com/"
        )
        
        st.markdown("---")
        
        st.header("📋 Features")
        st.markdown("""
        - ✅ 100% Data Capture
        - ✅ AI-Powered Structuring
        - ✅ Preserves Original Language
        - ✅ Contextual Comments
        - ✅ Formatted Excel Output
        """)
        
        st.markdown("---")
        
        st.header("ℹ️ How it Works")
        st.markdown("""
        1. **Upload PDF**: Choose your document
        2. **Extract Text**: PDF content is extracted
        3. **AI Analysis**: Claude identifies key-value pairs
        4. **Enhance Comments**: Adds contextual information
        5. **Download Excel**: Get your structured data
        """)
        
        st.markdown("---")
        st.caption("Powered by Claude AI & Streamlit")
    
    # Main content
    if not api_key:
        st.markdown('<div class="warning-box">⚠️ Please enter your Anthropic API key in the sidebar to get started.</div>', unsafe_allow_html=True)
        
        with st.expander("📖 How to get an API key"):
            st.markdown("""
            1. Visit [Anthropic Console](https://console.anthropic.com/)
            2. Sign up or log in to your account
            3. Navigate to API Keys section
            4. Create a new API key
            5. Copy and paste it in the sidebar
            """)
        return
    
    # File upload
    st.header("📤 Upload Your PDF")
    uploaded_file = st.file_uploader(
        "Choose a PDF file",
        type=['pdf'],
        help="Upload the unstructured PDF document you want to convert"
    )
    
    if uploaded_file is not None:
        # Display file info
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("File Name", uploaded_file.name)
        with col2:
            st.metric("File Size", f"{uploaded_file.size / 1024:.2f} KB")
        with col3:
            st.metric("File Type", uploaded_file.type)
        
        st.markdown("---")
        
        # Process button
        if st.button("🚀 Start Processing", type="primary", use_container_width=True):
            try:
                # Initialize extractor
                extractor = PDFToExcelExtractor(api_key)
                
                # Progress tracking
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                # Step 1: Extract text
                status_text.text("📄 Step 1/4: Extracting text from PDF...")
                progress_bar.progress(25)
                pdf_text = extractor.extract_text_from_pdf(uploaded_file)
                
                with st.expander("🔍 View Extracted Text (Preview)"):
                    st.text(pdf_text[:1000] + "..." if len(pdf_text) > 1000 else pdf_text)
                
                # Step 2: Structure data
                status_text.text("🤖 Step 2/4: AI is structuring your data...")
                progress_bar.progress(50)
                structured_data = extractor.structure_data_with_ai(pdf_text)
                
                # Step 3: Enhance comments
                status_text.text("✨ Step 3/4: Enhancing comments...")
                progress_bar.progress(75)
                structured_data = extractor.enhance_comments(structured_data, pdf_text)
                
                # Step 4: Create Excel
                status_text.text("📊 Step 4/4: Creating Excel file...")
                progress_bar.progress(90)
                excel_data = extractor.create_excel_output(structured_data)
                
                progress_bar.progress(100)
                status_text.text("✅ Processing Complete!")
                
                st.markdown("---")
                
                # Success message
                st.markdown('<div class="success-box">🎉 <strong>Success!</strong> Your document has been processed successfully.</div>', unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Display results
                st.header("📊 Extracted Data Preview")
                
                df = pd.DataFrame(structured_data)
                df.insert(0, '#', range(23, 23 + len(df)))
                
                st.dataframe(
                    df,
                    use_container_width=True,
                    height=400
                )
                
                # Statistics
                st.header("📈 Extraction Statistics")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Records", len(structured_data))
                with col2:
                    comments_count = sum(1 for item in structured_data if item.get('Comments', '').strip())
                    st.metric("Records with Comments", comments_count)
                with col3:
                    avg_comment_length = sum(len(item.get('Comments', '')) for item in structured_data) / len(structured_data)
                    st.metric("Avg Comment Length", f"{avg_comment_length:.0f} chars")
                with col4:
                    st.metric("Source Characters", len(pdf_text))
                
                st.markdown("---")
                
                # Download section
                st.header("💾 Download Results")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Download Excel
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_data,
                        file_name=f"Output_{timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col2:
                    # Download JSON
                    json_data = json.dumps(structured_data, indent=2)
                    st.download_button(
                        label="📥 Download JSON Data",
                        data=json_data,
                        file_name=f"Output_{timestamp}.json",
                        mime="application/json",
                        use_container_width=True
                    )
                
                # Sample records
                with st.expander("👁️ View Sample Records"):
                    for i, item in enumerate(structured_data[:5], 1):
                        st.markdown(f"**Record {i}:**")
                        st.write(f"**Key:** {item['Key']}")
                        st.write(f"**Value:** {item['Value']}")
                        st.write(f"**Comments:** {item.get('Comments', 'N/A')}")
                        st.markdown("---")
                
            except Exception as e:
                st.error(f"❌ Error during processing: {str(e)}")
                st.exception(e)
    
    else:
        st.markdown('<div class="info-box">👆 Please upload a PDF file to get started</div>', unsafe_allow_html=True)
        
        # Example/Demo section
        with st.expander("📚 See Example Output"):
            st.markdown("""
            ### Example Input (PDF):
            ```
            John passed 12th standard in 2007 with 92.50% score. His core subjects 
            included Mathematics, Physics, Chemistry, and Computer Science, 
            demonstrating his early aptitude for technical fields...
            ```
            
            ### Example Output (Excel):
            | # | Key | Value | Comments |
            |---|-----|-------|----------|
            | 23 | 12th standard pass out year | 2007 | His core subjects included Mathematics, Physics, Chemistry, and Computer Science... |
            | 24 | 12th overall board score | 92.50% | Outstanding achievement |
            """)


if __name__ == "__main__":
    main()
